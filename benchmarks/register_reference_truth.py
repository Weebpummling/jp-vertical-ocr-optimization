"""Register the human-verified academy dataset as ReferenceTruth and fix the split.

Reads the (local, never-committed) academy xlsx, mints a durable academy_id per
officer, assigns a deterministic train/hold-out split, and writes to the private data
home:

    groundtruth/split-manifest.csv   academy_id, cohort, branch, fullname_simp, use_flag
    groundtruth/registration.json    provenance + coverage map + manifest hash

Only the SHA-256 hashes and aggregate coverage go to the public repo
(docs/ground-truth-split.md). The split is HASH-DETERMINISTIC: an officer's
assignment depends only on their identity key, never on row order, file version, or a
random seed — so the split is reproducible from the data alone and cannot drift.

Rule: holdout if int(sha1(academy_id + SALT), 16) % 100 < HOLDOUT_PCT.
Expected ~30% per (cohort x branch) stratum by the law of large numbers; exact
proportions per stratum vary and are reported in the coverage map.

Hold-out discipline (docs/PLAN.md): hold-out rows are never used for VLM tuning,
prompt iteration, threshold fitting, or error analysis. Enforcement owner: the lead.

Usage:
    python register_reference_truth.py <cohorts.xlsx> <data_home_groundtruth_dir>
"""
import csv
import hashlib
import json
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl

HOLDOUT_PCT = 30
SALT = "jp-vertical-ocr-v1"   # frozen; changing it redefines the split


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def main(xlsx: Path, outdir: Path) -> None:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(rows)]
    ix = {name: header.index(name) for name in
          ("cohort", "branch", "fullname", "fullname_simp")}

    seen: Counter = Counter()
    records = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue
        cohort = str(row[ix["cohort"]]).strip()
        branch = str(row[ix["branch"]]).strip()
        fullname = str(row[ix["fullname"]]).strip()
        simp = str(row[ix["fullname_simp"]] or "").strip()
        key = f"{cohort}|{fullname}|{branch}"
        seq = seen[key]           # disambiguate true duplicates deterministically
        seen[key] += 1
        academy_id = hashlib.sha1(f"{key}|{seq}".encode("utf-8")).hexdigest()[:16]
        flag = ("holdout"
                if int(hashlib.sha1((academy_id + SALT).encode()).hexdigest(), 16) % 100
                < HOLDOUT_PCT else "train")
        records.append((academy_id, cohort, branch, simp or fullname, flag))

    dupes = {k: c for k, c in seen.items() if c > 1}

    outdir.mkdir(parents=True, exist_ok=True)
    manifest = outdir / "split-manifest.csv"
    with open(manifest, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["academy_id", "cohort", "branch", "fullname_simp", "use_flag"])
        w.writerows(records)

    flags = Counter(r[4] for r in records)
    per_cohort: dict = defaultdict(Counter)
    for _, cohort, _, _, flag in records:
        per_cohort[cohort][flag] += 1
    coverage = {c: {"train": n["train"], "holdout": n["holdout"]}
                for c, n in sorted(per_cohort.items(), key=lambda kv: int(kv[0]))}

    reg = {
        "asset": "ReferenceTruth (academy dataset, human-verified)",
        "registered": date.today().isoformat(),
        "source_file": xlsx.name,
        "source_sha256": sha256_file(xlsx),
        "n_records": len(records),
        "n_duplicate_keys": len(dupes),
        "split_rule": f"holdout if sha1(id+'{SALT}') % 100 < {HOLDOUT_PCT}",
        "flags": dict(flags),
        "manifest_sha256": sha256_file(manifest),
        "coverage_by_cohort": coverage,
    }
    with open(outdir / "registration.json", "w", encoding="utf-8") as f:
        json.dump(reg, f, indent=1, ensure_ascii=False)

    print(json.dumps({k: v for k, v in reg.items() if k != "coverage_by_cohort"},
                     indent=1))
    print("cohorts:", len(coverage), "| duplicate identity keys:", len(dupes))
    if dupes:
        for k, c in list(dupes.items())[:5]:
            print("  dup:", k, "x", c)


if __name__ == "__main__":
    main(Path(sys.argv[1]), Path(sys.argv[2]))
